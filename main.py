from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from tkinter import *
from openpyxl import *

gauth = GoogleAuth()
gauth.LocalWebserverAuth()
drive = GoogleDrive(gauth)

goodwb = load_workbook('teams.xlsx')
goodSheet = goodwb.active
def goodExcel():
    goodSheet.column_dimensions['A'].width = 10
    goodSheet.column_dimensions['B'].width = 10
    goodSheet.column_dimensions['C'].width = 10
    goodSheet.column_dimensions['D'].width = 10
    goodSheet.column_dimensions['E'].width = 20
    goodSheet.column_dimensions['F'].width = 20
    goodSheet.cell(row=1, column=1).value = "Team #"
    goodSheet.cell(row=1, column=2).value = "ballsScoredLower"
    goodSheet.cell(row=1, column=3).value = "ballsScoredUpper"
    goodSheet.cell(row=1, column=4).value = "Highest Climb Level"
    goodSheet.cell(row=1, column=5).value = "School"
    goodSheet.cell(row=1, column=6).value = "Student or Parent Led"
def focus1(event):
    ballsScoredLower.focus_set()
def focus2(event):
    ballsScoredUpper.focus_set()
def focus3(event):
    climbLevel.focus_set()
def focus4(event):
    school.focus_set()
def focus5(event):
    studentOrParentLed.focus_set()
def focus6(event):
    badOrGood.focus_set()
def clear():
    teamNum.delete(0, END)
    ballsScoredLower.delete(0, END)
    ballsScoredUpper.delete(0, END)
    climbLevel.delete(0, END)
    school.delete(0, END)
    studentOrParentLed.delete(0, END)
    badOrGood.delete(0, END)
def sheet():
    if (teamNum.get() == "" and
        ballsScoredLower.get() == "" and
        ballsScoredUpper.get() == "" and
        climbLevel.get() == "" and
        school.get() == "" and
        studentOrParentLed.get() == ""):
        print("empty input")
    else:
        current_row = goodSheet.max_row
        current_column = goodSheet.max_column
        goodSheet.cell(row=current_row + 1, column=1).value = teamNum.get()
        goodSheet.cell(row=current_row + 1, column=2).value = lowerCount
        goodSheet.cell(row=current_row + 1, column=3).value = upperCount
        goodSheet.cell(row=current_row + 1, column=4).value = climbLevel.get()
        goodSheet.cell(row=current_row + 1, column=5).value = school.get()
        goodSheet.cell(row=current_row + 1, column=6).value = studentOrParentLed.get()
        goodwb.save('teams.xlsx')
        teamNum.focus_set()
        clear()
def upperButtonPress():
    global upperCount
    upperCount = 1 + upperCount
def lowerButtonPress():
    global lowerCount
    lowerCount = 1 + lowerCount
def updateFile():
    file5 = drive.CreateFile()
    # Read file and set it as a content of this instance.
    file5.SetContentFile('teams.xlsx')
    file5.Upload() # Upload the file.

if __name__ == "__main__":
    lowerCount = 0
    upperCount = 0
    root = Tk()
    root.configure(background='blue')
    root.title("FRC Scouting")
    # set the configuration of GUI window
    root.geometry("400x700")
    goodExcel()
    heading = Label(root, text="Scouting Form", bg="white")
    num= Label(root, text="Team Number", bg="white")
    lower = Label(root, text="Balls Scored Lower", bg="White")
    upper = Label(root, text="Balls Scored Upper", bg="White")
    climb = Label(root, text="Climb Level", bg="White")
    whatSchool = Label(root, text="School", bg="White")
    studentOrParent = Label(root, text="Student Or Parent Led", bg="White")
    theDecision = Label(root, text="Save after each match", bg="white")
    heading.grid(row=0, column=0)
    num.grid(row=1, column=0)
    lower.grid(row=3, column=0)
    upper.grid(row=5, column=0)
    climb.grid(row=7, column=0)
    whatSchool.grid(row=9, column=0)
    studentOrParent.grid(row=11, column=0)
    theDecision.grid(row=13, column=0)
    teamNum = Entry(root)
    ballsScoredLower = Entry(root)
    ballsScoredUpper = Entry(root)
    climbLevel = Entry(root)
    school = Entry(root)
    studentOrParentLed = Entry(root)
    badOrGood = Entry(root) 
    teamNum.bind("<Return>", focus1)
    ballsScoredLower.bind("<Return>", focus2)
    ballsScoredUpper.bind("<Return>", focus3)
    climbLevel.bind("<Return>", focus4)
    school.bind("<Return>", focus5)
    studentOrParentLed.bind("<Return>", focus6)
    teamNum.grid(row=2, column=0, ipadx="25")
    lowerButton = Button(root, text="+1 Lower", fg="White", bg="Black", command=lowerButtonPress)
    lowerButton.grid(row=4, column=0, ipadx="25")
    upperButton = Button(root, text="+1 Upper", fg="White", bg="Black", command=upperButtonPress)
    upperButton.grid(row=6, column=0, ipadx="25")
    climbLevel.grid(row=8, column=0, ipadx="25")
    school.grid(row=10, column=0, ipadx="25")
    studentOrParentLed.grid(row=12, column=0, ipadx="25") 
    goodExcel()
    good = Button(root, text="Save", fg="Black", bg="Green", command=sheet)
    good.grid(row=15, column=0)
    upload = Button(root, text="upload to drive", fg="light blue", bg="Black", command=updateFile)
    upload.grid(row=16, column=0)
    root.mainloop()
