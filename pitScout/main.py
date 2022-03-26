from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from tkinter import *
from openpyxl import *

gauth = GoogleAuth()
drive = GoogleDrive(gauth)

goodwb = load_workbook('pitTeams.xlsx')
goodSheet = goodwb.active
def goodExcel():
    goodSheet.column_dimensions['A'].width = 10
    goodSheet.column_dimensions['B'].width = 10
    goodSheet.column_dimensions['C'].width = 10
    goodSheet.column_dimensions['D'].width = 10
    goodSheet.column_dimensions['E'].width = 20
    goodSheet.column_dimensions['F'].width = 20
    goodSheet.cell(row=1, column=1).value = "Team #"
    goodSheet.cell(row=1, column=2).value = "driveTrain"
    goodSheet.cell(row=1, column=3).value = "ballsScoredUpper"
    goodSheet.cell(row=1, column=4).value = "Highest Climb Level"
    goodSheet.cell(row= 1, column=7).value = "Autonomous y/n"
    goodSheet.cell(row=1, column=9).value = "Stability (1-10) 1 = getting flipped"
def focus1(event):
    drive.focus_set()
def focus2(event):
    studentOrDad.focus_set()
def focus3(event):
    robotType.focus_set()
def focus4(event):
    whatSchool.focus_set()
def focus5(event):
    contactName.focus_set()
def focus6(event):
    teamComments.focus_set()
def clear():
    teamNum.delete(0, END)
    drive.delete(0, END)
    studentOrDad.delete(0, END)
    robotType.delete(0, END)
    whatSchool.delete(0, END)
    contactName.delete(0, END)
    teamComments.delete(0, END)
def sheet():
    if (teamNum.get() == "" and
        drive.get() == "" and
        studentOrDad.get() == "" and
        robotType.get() == "" and
        whatSchool.get() == "" and
        contactName.get() == ""):
        print("empty input")
    else:
        current_row = goodSheet.max_row
        current_column = goodSheet.max_column
        goodSheet.cell(row=current_row + 1, column=1).value = teamNum.get()
        goodSheet.cell(row=current_row + 1, column=2).value = drive.get()
        goodSheet.cell(row=current_row + 1, column=3).value = studentOrDad.get()
        goodSheet.cell(row=current_row + 1, column=4).value = robotType.get()
        goodSheet.cell(row=current_row + 1, column=5).value = whatSchool.get()
        goodSheet.cell(row=current_row + 1, column=6).value = contactName.get()
        goodwb.save('pitTeams.xlsx')
        teamNum.focus_set()
        clear()
def updateFile():
    gauth.LocalWebserverAuth()
    file5 = drive.CreateFile()
    # Read file and set it as a content of this instance.
    file5.SetContentFile('pitTeams.xlsx')
    file5.Upload() # Upload the file.

if __name__ == "__main__":
    root = Tk()
    root.configure(background='blue')
    root.title("FRC Pit Scouting")
    # set the configuration of GUI window
    root.geometry("400x700")
    goodExcel()
    heading = Label(root, text="Pit Scouting Form", bg="white")
    num= Label(root, text="Team Number", bg="white")
    driveTrain = Label(root, text="Drive Train", bg="White")
    studentLed = Label(root, text="Student or Dad team", bg="White")
    type = Label(root, text="Kind of bot (climb = 1 shoot = 2 both = 3 defense(potato) = 0)", bg="White")
    school = Label(root, text="What school", bg="White")
    contact = Label(root, text="Contact person name", bg="White")
    comments = Label(root, text="Team's Comments", bg="White")
    save = Label(root, text="Save after each match", bg="white")
    heading.grid(row=0, column=0)
    num.grid(row=1, column=0)
    driveTrain.grid(row=3, column=0)
    studentLed.grid(row=5, column=0)
    type.grid(row=7, column=0)
    school.grid(row=9, column=0)
    contact.grid(row=11, column=0)
    comments.grid(row=13, column=0)
    save.grid(row=15, column=0)
    teamNum = Entry(root)
    drive = Entry(root)
    studentOrDad = Entry(root)
    robotType = Entry(root)
    whatSchool = Entry(root)
    contactName = Entry(root)
    teamComments = Entry(root)
    teamNum.bind("<Return>", focus1)
    drive.bind("<Return>", focus2)
    studentLed.bind("<Return>", focus3)
    robotType.bind("<Return>", focus4)
    contactName.bind("<Return>", focus5)
    teamComments.bind("<Return>", focus6)
    teamNum.grid(row=2, column=0, ipadx="25")
    drive.grid(row=4, column=0, ipadx="25")
    studentOrDad.grid(row=6, column=0, ipadx="25")
    robotType.grid(row=8, column=0, ipadx="25")
    whatSchool.grid(row=10, column=0, ipadx="25")   
    contactName.grid(row=12, column=0, ipadx="25")
    teamComments.grid(row=14, column=0, ipadx="25") 
    goodExcel()
    good = Button(root, text="Save", fg="Black", bg="Green", command=sheet)
    good.grid(row=15, column=0)
    upload = Button(root, text="upload to drive", fg="light blue", bg="Black", command=updateFile)
    upload.grid(row=16, column=0)
    root.mainloop()
